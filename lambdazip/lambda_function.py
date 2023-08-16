import boto3
import openpyxl
from botocore.exceptions import ClientError
# import datetime
import pytz
from pytz import timezone
from tabulate import tabulate
import io
# import datetime
import os
from datetime import datetime

def update_excel_with_instance_details(event, context):
    # Configure Boto3 clients
    s3_client = boto3.client('s3')
    ec2_client = boto3.client('ec2')
    elbv2_client = boto3.client('elbv2')
    
    # Specify the S3 bucket and file name
    bucket_name = os.environ['bucket_name']
    file_name = os.environ['file_name']
    account = os.environ['account']
    region = os.environ['region']
    
    
    
    # Download the Excel file from S3
    excel_file = "/tmp/excel_file.xlsx"
    s3_client.download_file(bucket_name, file_name, excel_file)
    
    # Table List
    
    table_list = ''
    
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['EC2']
    
    # Find the column index of instance IDs (assuming it's the first column)
    instance_id_column = 1
    
    # Loop through rows and get instance IDs
    status_list = [['Instatnce Id','Tags', 'Instance Status','Instance Status Check' ,'System Status Check','Public IP', 'Private IpAddress', 'Instance Type' , 'Availability Zone' , 'Key Name' ]]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=instance_id_column, max_col=instance_id_column):
        instance_id = row[0].value
        if instance_id:
            try:
                instance_state = 'Instance Not Found'  # Initialize instance_state with a default value
                Public_Ip = '-'
                PrivateIpAddress = '-'
                InstanceType = '-'
                AvailabilityZone = '-'
                KeyName = '-'
                Tags = '-'
                instance_status_check = '-'
                system_status_check = '-'
                
                # Check instance status
                response = ec2_client.describe_instances(InstanceIds=[instance_id])
                response2 = ec2_client.describe_instance_status(InstanceIds=[instance_id])
                if 'Reservations' in response and len(response['Reservations']) > 0:
                    instance = response['Reservations'][0]['Instances'][0]
                    print("Data:", instance)
                    instance_state = instance.get('State', {}).get('Name', 'Instance Not Found')
                    Public_Ip = instance.get('PublicIpAddress', '-')
                    PrivateIpAddress= instance.get('PrivateIpAddress', '-')
                    InstanceType= instance.get('InstanceType', '-')
                    AvailabilityZone= instance.get('Placement', {}).get('AvailabilityZone', '-')
                    KeyName= instance.get('KeyName', '-')
                    Tags=instance['Tags'][0]['Value']
                    # instance_status_check = response2['InstanceStatuses'][0]['InstanceStatus']['Status']
                    # system_status_check = response2['InstanceStatuses'][0]['SystemStatus']['Status']
                    if 'InstanceStatuses' in response2 and len(response2['InstanceStatuses']) > 0:
                         instance_status_check = response2['InstanceStatuses'][0]['InstanceStatus']['Status']
                         system_status_check = response2['InstanceStatuses'][0]['SystemStatus']['Status']
                    else:
                         instance_status_check = 'No instance status found'
                         system_status_check = 'No system status found'
                    
                    
                status_list.append([instance_id, Tags, instance_state ,instance_status_check,system_status_check, Public_Ip, PrivateIpAddress, InstanceType, AvailabilityZone, KeyName ])  # Changed to list format
            
                
                sheet.cell(row=row[0].row, column= 2, value=Tags )
                sheet.cell(row=row[0].row, column= 3, value=instance_state )
                sheet.cell(row=row[0].row, column= 4, value=instance_status_check )
                sheet.cell(row=row[0].row, column= 5, value=system_status_check )
                sheet.cell(row=row[0].row, column= 6, value=Public_Ip )
                sheet.cell(row=row[0].row, column= 7, value=PrivateIpAddress )
                sheet.cell(row=row[0].row, column= 8, value=InstanceType )
                sheet.cell(row=row[0].row, column= 8, value=AvailabilityZone )
                sheet.cell(row=row[0].row, column= 10, value=KeyName )
                
                
            except ClientError as e:
                
                    
                    
                    instance_state = 'Instance Not Found'  # Initialize instance_state with a default value
                    Public_Ip = '-'
                    PrivateIpAddress = '-'
                    InstanceType = '-'
                    AvailabilityZone = '-'
                    KeyName = '-'
                    Tags = '-'
                    instance_status_check = '-'
                    system_status_check = '-'
                    
                    status_list.append([instance_id, Tags, instance_state ,instance_status_check,system_status_check, Public_Ip, PrivateIpAddress, InstanceType, AvailabilityZone, KeyName ])  # Changed to list format
                    
                    sheet.cell(row=row[0].row, column= 2, value=Tags )
                    sheet.cell(row=row[0].row, column= 3, value=instance_state )
                    sheet.cell(row=row[0].row, column= 4, value=instance_status_check )
                    sheet.cell(row=row[0].row, column= 5, value=system_status_check )
                    sheet.cell(row=row[0].row, column= 6, value=Public_Ip )
                    sheet.cell(row=row[0].row, column= 7, value=PrivateIpAddress )
                    sheet.cell(row=row[0].row, column= 8, value=InstanceType )
                    sheet.cell(row=row[0].row, column= 8, value=AvailabilityZone )
                    sheet.cell(row=row[0].row, column= 10, value=KeyName )
                    
                    
                    
            
        else:
            print("Empty cell encountered in instance ID column")
            
            
    # Save Excel file in memory
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    
    ec2_table_html ='<div style="margin-bottom: 20px;">'
    ec2_table_html +='<h3>EC2 Inventor</h3>'
    ec2_table_html += "<table border='1'><tr>"
    headers = status_list[0]
    for header in headers:
        ec2_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    ec2_table_html += "</tr>"
    
    
    for row in status_list[1:]:
        ec2_table_html += "<tr>"
        for cell in row:
            ec2_table_html += f'<td style="background-color: {"rgb(205, 86, 86)"if cell in ("Instance Not Found", "stopped" , "stopping", "terminated", "terminating", "shutting-down", "impaired", "failed" ) else "white"};">{cell}</td>'
        ec2_table_html += "</tr>"
    
    ec2_table_html += "</table>"
    ec2_table_html += '</div>'
    
    


#########################################################################################################################################
##                                               Code for ELB
########################################################################################################################################

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['ELB']
    
    # Find the column index of instance IDs (assuming it's the first column)
    elb_column = 1
    
    # Loop through rows and get instance IDs
    elb_status_list = [['LoadBalancer Name','State', 'Scheme', 'Type', 'Vpc Id' ]]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=elb_column, max_col=elb_column):
        elb_name = row[0].value
        if elb_name:
            try:
                  # Initialize instance_state with a default value
                State = 'LB Not Found'
                
                Type = '-'
                Scheme = '-'
                VpcId = '-'
                
                # Check instance status
                response = elbv2_client.describe_load_balancers(Names=[elb_name])
                if 'LoadBalancers' in response and len(response['LoadBalancers']) > 0:
                    elb = response['LoadBalancers'][0]
                    
                    
                    State= elb.get('State', {}).get('Code', 'LB Not Found')
                    
                    Scheme = elb.get('Scheme', '-')
                    Type = elb.get('Type', '-')
                    VpcId = elb.get('VpcId', '-')
                    
                    
                elb_status_list.append([ elb_name, State ,Scheme,Type, VpcId  ])  # Changed to list format
                
                
                sheet.cell(row=row[0].row, column= 3, value=State )
                sheet.cell(row=row[0].row, column= 4, value=Scheme )
                sheet.cell(row=row[0].row, column= 5, value=Type )
                sheet.cell(row=row[0].row, column= 6, value=VpcId )
                
            except ClientError as e:
                    print(e)
                     # Initialize instance_state with a default value
                    State = 'LB Not Found'
                    
                    Type = '-'
                    Scheme = '-'
                    VpcId = '-'
                    
                    elb_status_list.append([ elb_name, State ,Scheme,Type, VpcId  ]) 
                    
                    
                    sheet.cell(row=row[0].row, column= 3, value=State )
                    sheet.cell(row=row[0].row, column= 4, value=Scheme )
                    sheet.cell(row=row[0].row, column= 5, value=Type )
                    sheet.cell(row=row[0].row, column= 6, value=VpcId )
                    
                    
                
        else:
            print("Empty cell encountered in ELB ARN column")
            
            
     # Save Excel file in memory
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    elb_table_html ='<div style="margin-bottom: 20px;">'
    elb_table_html +='<h3>ELB Inventory</h3>'
    elb_table_html += "<table border='1'><tr>"
    headers = elb_status_list[0]
    for header in headers:
        elb_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    elb_table_html += "</tr>"
    
    
    for row in elb_status_list[1:]:
        elb_table_html += "<tr>"
        for cell in row:
            elb_table_html += f'<td style="background-color: {"rgb(205, 86, 86)"if cell in ("LB Not Found", "outofservice" , "unknown", "failed" ) else "white"};">{cell}</td>'
        elb_table_html += "</tr>"
    
    elb_table_html += "</table>"
    elb_table_html += '</div>'
    

    

##########################################################################################################################################
#                                                              Volumes
######################################################################################################################################

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['Volume']
    
    # Find the column index of instance IDs (assuming it's the first column)
    vol_column = 1
    
    # Loop through rows and get instance IDs
    vol_status_list = [['Volume Id', 'Volume Size', 'Volume State', 'Volume Type','Volume AZ', 'Volume Attachment', 'Encryption', 'Creation Time']]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=vol_column, max_col=vol_column):
        vol_id = row[0].value
        if vol_id:
            
                # Volume_id = 'N/A'  # Initialize instance_state with a default value
                # Volume_Name = 'N/A' 
                Volume_Size = '-' 
                Volume_State  = '-' 
                Volume_AZ = '-' 
                Volume_Attachment = '-' 
                Encryption = '-' 
                Creation_Time = '-' 
                Volume_Type = '-'
    
                # Check instance status
                
                
                # Create an EC2 resource object
                ec2_client = boto3.resource('ec2')
                
                # Get all volumes
        try:
                responce = ec2_client.volumes.filter(VolumeIds=[vol_id])
                for volume in responce:
                              Volume_id=volume.id
                              Volume_Size=volume.size
                              Volume_Type=volume.volume_type
                              Volume_State= f"{volume.state}"
                              Creation_Time=f"{volume.create_time}"
                              Volume_AZ=volume.availability_zone
                              Volume_State= volume.state
                              Encryption=volume.encrypted
                            #   if volume is not None:
                            #             name_tag = next((tag['Value'] for tag in volume.tags if tag['Key'] == 'Name'), 'N/A')
                            # #   name_tag = next((tag['Value'] for tag in volume.tags if tag['Key'] == 'Name'), 'N/A')
                            
                              
                              
                              vol_status_list.append([Volume_id, Volume_Size, Volume_State, Volume_Type, Volume_AZ, Volume_Attachment, Encryption, Creation_Time])  # Changed to list format                  
                              
                              sheet.cell(row=row[0].row, column= 2, value=Volume_Size )
                              sheet.cell(row=row[0].row, column= 3, value=Volume_State )
                              sheet.cell(row=row[0].row, column= 4, value=Volume_Type )
                              sheet.cell(row=row[0].row, column= 5, value=Volume_AZ )
                              sheet.cell(row=row[0].row, column= 6, value=Volume_Attachment )
                              sheet.cell(row=row[0].row, column= 7, value=Encryption )
                              sheet.cell(row=row[0].row, column= 8, value=Creation_Time )
                
                              
        except ClientError as e:
                   
                    
                    Volume_id=vol_id
                    Volume_Size = '-' 
                    Volume_State  = 'Volume Not Found' 
                    Volume_AZ = '-' 
                    Volume_Attachment = '-'
                    Encryption = '-'
                    Creation_Time = '-'
                    Volume_Type = '-'
                    
                    vol_status_list.append([Volume_id, Volume_Size, Volume_State, Volume_Type, Volume_AZ, Volume_Attachment, Encryption, Creation_Time])
                    
                    # sheet.cell(row=row[0].row, column= 2, value=Volume_Name )
                    sheet.cell(row=row[0].row, column= 2, value=Volume_Size )
                    sheet.cell(row=row[0].row, column= 3, value=Volume_State )
                    sheet.cell(row=row[0].row, column= 4, value=Volume_Type )
                    sheet.cell(row=row[0].row, column= 5, value=Volume_AZ )
                    sheet.cell(row=row[0].row, column= 6, value=Volume_Attachment )
                    sheet.cell(row=row[0].row, column= 7, value=Encryption )
                    sheet.cell(row=row[0].row, column= 8, value=Creation_Time )

        else:
            print("Empty cell encountered in Volunme ID column")
            

            

    # Save Excel file in memory
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    
    vol_table_html ='<div style="margin-bottom: 20px;">'
    vol_table_html +='<h3>Volume Inventory</h3>'
    vol_table_html += "<table border='1'><tr>"
    headers = vol_status_list[0]
    for header in headers:
        vol_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    vol_table_html += "</tr>"
    
    
    for row in vol_status_list[1:]:
        vol_table_html += "<tr>"
        for cell in row:
            vol_table_html += f'<td style="background-color: {"rgb(205, 86, 86)"if cell == "Volume Not Found" else "white"};">{cell}</td>'
        vol_table_html += "</tr>"
    
    vol_table_html += "</table>"
    vol_table_html += '</div>'
    


##################################################################################################################
#                                                  Code for IP
##################################################################################################################

    

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['EIP']
    
    # Find the column index of instance IDs (assuming it's the first column)
    eip_id_column = 1
    
    # Loop through rows and get instance IDs
    eip_list = [['Public IP','Eip Name', 'Type', 'Associated instance ID', 'Private IP' ]]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=eip_id_column, max_col=eip_id_column):
        eip_address = row[0].value
        

        try:
            ec2_client = boto3.client('ec2')
            response = ec2_client.describe_addresses(PublicIps=[eip_address])
            eip_info = response['Addresses'][0] if response['Addresses'] else {}
            
            Eip_name = '-'
            tags = eip_info.get('Tags', [])
            for tag in tags:
                  if tag['Key'] == 'Name':
                      Eip_name = tag['Value']
                  break 
            
            
            # Public_IP = eip_info.get('PublicIp')
            Eip_type = eip_info.get('Type', '-')
            Associated_instance_id = eip_info.get('InstanceId', '-')
            Private_IP = eip_info.get('PrivateIpAddress', '-') 
            
            eip_list.append([eip_address,Eip_name, Eip_type, Associated_instance_id, Private_IP])
            
            
            
            sheet.cell(row=row[0].row, column= 2, value=Eip_name )
            sheet.cell(row=row[0].row, column= 3, value=Eip_type )
            sheet.cell(row=row[0].row, column= 4, value=Associated_instance_id )
            sheet.cell(row=row[0].row, column= 5, value=Private_IP )
            
        except ClientError as e:
            #  print("An error occurred:", e)
             
            
                Eip_name = 'EIP Not Found'
                Eip_type = '-'
                Associated_instance_id = '-'
                Private_IP = '-'
                
                eip_list.append([eip_address,Eip_name, Eip_type, Associated_instance_id, Private_IP])
                
                
            
                sheet.cell(row=row[0].row, column= 2, value=Eip_name )
                sheet.cell(row=row[0].row, column= 3, value=Eip_type )
                sheet.cell(row=row[0].row, column= 4, value=Associated_instance_id )
                sheet.cell(row=row[0].row, column= 5, value=Private_IP )
            
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    
    eip_table_html ='<div style="margin-bottom: 20px;">'
    eip_table_html +='<h3>EIP Inventory</h3>'
    eip_table_html += "<table border='1'><tr>"
    headers = eip_list[0]
    for header in headers:
        eip_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    eip_table_html += "</tr>"
    
    
    for row in eip_list[1:]:
        eip_table_html += "<tr>"
        for cell in row:
            eip_table_html += f'<td style="background-color: {"rgb(205, 86, 86)" if cell == "EIP Not Found" else "white"};">{cell}</td>'
        eip_table_html += "</tr>"
    
    eip_table_html += "</table>"
    eip_table_html += '</div>'
    

            
 ############################################## Code for S3 Bucket ########################################################
 
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['S3']
    
    # Find the column index of instance IDs (assuming it's the first column)
    s3_id_column = 1
    
    # Loop through rows and get Bucket Name
    s3_list = [['Bucket Name','Bucket Region', 'Creation Date' ]]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=s3_id_column, max_col=s3_id_column):
        s3_bucket = row[0].value
        
        
         

        try:
            bucket_response = s3_client.head_bucket(Bucket=s3_bucket)
            creation_date = bucket_response['ResponseMetadata']['HTTPHeaders']['date']
            creation_region = bucket_response['ResponseMetadata']['HTTPHeaders']['x-amz-bucket-region']
            
            creation_date= str(creation_date)
            
            sheet.cell(row=row[0].row, column= 2, value=creation_region )
            sheet.cell(row=row[0].row, column= 3, value=creation_date )
            
            s3_list.append([s3_bucket,creation_region,creation_date ])
            
        except ClientError as e:
            #  print("An error occurred:", e)
             
            
                creation_date = '-'
                creation_region = 'Bucket Not Found'
                
                s3_list.append([s3_bucket,creation_region,creation_date ])
                
                
            
                sheet.cell(row=row[0].row, column= 2, value=creation_region )
                sheet.cell(row=row[0].row, column= 3, value=creation_date )
             
            
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    s3_table_html ='<div style="margin-bottom: 20px;">'
    s3_table_html +='<h3>S3 Bucket Inventory</h3>'
    s3_table_html += "<table border='1'><tr>"
    headers = s3_list[0]
    for header in headers:
        s3_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    s3_table_html += "</tr>"
    
    
    for row in s3_list[1:]:
        s3_table_html += "<tr>"
        for cell in row:
            s3_table_html += f'<td style="background-color: {"rgb(205, 86, 86)"if cell == "Bucket Not Found" else "white"};">{cell}</td>'
        s3_table_html += "</tr>"
    
    s3_table_html += "</table>"     
    s3_table_html += '</div>'    
            
            
    
            
################################################## Code for Key Pair ##########################################################

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['KeyPair']
    
    # Find the column index of instance IDs (assuming it's the first column)
    key_name_column = 1
    
    # Loop through rows and get Bucket Name
    key_list = [['Key name','Key Type', 'Creation Date', 'KeyPair ID' ]]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=key_name_column, max_col=key_name_column):
        key_name = row[0].value

        try:
            ec2 = boto3.resource('ec2')
            key_pair_info = ec2.KeyPair(key_name)
            
            key_type= key_pair_info.key_type
            key_create_time= f"{key_pair_info.create_time}"
            key_id= key_pair_info.key_pair_id
        
            key_list.append([key_name, key_type,key_create_time, key_id  ])
            
            
            sheet.cell(row=row[0].row, column= 4, value=key_create_time )
            sheet.cell(row=row[0].row, column= 5, value=key_id )
            
        except ClientError as e:
            #  print("An error occurred:", e)
             
            
                key_type= "Key Not Found"
                key_create_time= "-"
                key_id= "-"
                key_format= "-"
                
                key_list.append([key_name, key_type,key_create_time, key_id  ]) 
                
                
            
                sheet.cell(row=row[0].row, column= 2, value=key_type )
                sheet.cell(row=row[0].row, column= 3, value=key_format )
                sheet.cell(row=row[0].row, column= 4, value=key_create_time )
                sheet.cell(row=row[0].row, column= 5, value=key_id )
            
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    key_table_html ='<div style="margin-bottom: 20px;">'
    key_table_html +='<h3>Key Pair Inventory</h3>'

    key_table_html += "<table border='1'><tr>"
    headers = key_list[0]
    for header in headers:
        key_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    key_table_html += "</tr>"
    
    
    for row in key_list[1:]:
        key_table_html += "<tr>"
        for cell in row:
            key_table_html += f'<td style="background-color: {"rgb(205, 86, 86)"if cell == "Key Not Found" else "white"};">{cell}</td>'
        key_table_html += "</tr>"
    
    key_table_html += "</table>"  
    key_table_html += '</div>'      
    
    

###################################### Code for IAM User ##########################################################

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['IAM_User']
    
    # Find the column index of instance IDs (assuming it's the first column)
    user_column = 1
    
    # Loop through rows and get Bucket Name
    user_list = [['User Name', 'Groups', 'Last Activity', 'MFA', 'Password Age', 'Active Key Age' ]]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=user_column, max_col=user_column):
        user_name = row[0].value

        try:
            iam_client = boto3.client('iam')
            
            # Use the get_user method to describe the user
            response = iam_client.get_user(UserName=user_name)
            user_details = response['User']
            
            group_names = '-'
            password_last_used = '-'
            mfa_devices = '-'
            password_age = '-'
            
            # Get user's groups
            user_groups = iam_client.list_groups_for_user(UserName=user_name)['Groups']
            group_names = [group['GroupName'] for group in user_groups]
            group_names_str = ', '.join(group_names)
            
            # Get user's access keys
            access_keys = iam_client.list_access_keys(UserName=user_name)['AccessKeyMetadata']
            active_key_age = None
            if access_keys:
                oldest_key_creation = min(access_key['CreateDate'] for access_key in access_keys)
                active_key_age = (datetime.now() - oldest_key_creation.replace(tzinfo=None)).days
            
            # Get user's MFA devices
            mfa_devices_data = iam_client.list_mfa_devices(UserName=user_name)
            mfa_devices_list = mfa_devices_data.get('MFADevices', [])  # Using empty list if 'MFADevices' key is not present
            
            mfa_devices_value = len(mfa_devices_list)
            
            # Get user's password last used
            password_last_used = user_details.get('PasswordLastUsed')

            if password_last_used is not None:
                password_last_used = password_last_used.replace(tzinfo=None)
                password_age_timedelta = datetime.now() - password_last_used
                password_age_days = password_age_timedelta.days
                password_age_str = str(password_age_days)
            else:
                password_age_str = '-'
            
            password_age = (datetime.now() - user_details['CreateDate'].replace(tzinfo=None)).days
                        
                    
            user_list.append([user_name,group_names,password_last_used, mfa_devices_value,password_age, active_key_age])
            
            
            sheet.cell(row=row[0].row, column= 2, value=group_names_str )
            sheet.cell(row=row[0].row, column= 3, value=password_last_used )
            sheet.cell(row=row[0].row, column= 4, value=mfa_devices_value )
            sheet.cell(row=row[0].row, column= 5, value=password_age )
            sheet.cell(row=row[0].row, column= 6, value=active_key_age )
            
            
        except ClientError as e:
              
                print("\nNo such user found")
                
                group_names_str= "User Not Found"
                active_key_age= "-"
                mfa_devices="-"
                password_last_used= "-"
                password_age = "-"
                
                user_list.append([user_name,group_names_str,password_last_used, mfa_devices_value, password_age, active_key_age])
                
            
                sheet.cell(row=row[0].row, column= 2, value=group_names_str )
                sheet.cell(row=row[0].row, column= 3, value=password_last_used )
                sheet.cell(row=row[0].row, column= 4, value=mfa_devices_value )
                sheet.cell(row=row[0].row, column= 5, value=password_age )
                sheet.cell(row=row[0].row, column= 6, value=active_key_age )
        
              
            
                
                
            
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    
    user_table_html ='<div style="margin-bottom: 20px;">'
    user_table_html +='<h3>IAM User Inventory</h3>'
    user_table_html += "<table border='1'><tr>"
    headers = user_list[0]
    for header in headers:
        user_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    user_table_html += "</tr>"
    
    
    for row in user_list[1:]:
        user_table_html += "<tr>"
        for cell in row:
            user_table_html += f'<td style="background-color: {"rgb(205, 86, 86)"if cell == "User Not Found" else "white"};">{cell}</td>'
        user_table_html += "</tr>"
    
    user_table_html += "</table>"
    user_table_html += '</div>'
    
    print(user_table_html)
    
    
    
    ##################################################### Code for RDS #################################################################
    
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['RDS']
    
    # Find the column index of instance IDs (assuming it's the first column)
    rds_column = 1
    
    # Loop through rows and get Bucket Name
    rds_list = [['RDS ID', 'Size','Availability Zone', 'DBInstance Status', 'Engine Version', 'Endpoint', 'Multi AZ' ]]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=rds_column, max_col=rds_column):
        rds_id = row[0].value

        try:
            rds_client = boto3.client('rds')
            
            response = rds_client.describe_db_instances(DBInstanceIdentifier=rds_id)


            if 'DBInstances' in response and len(response['DBInstances']) > 0:
                db_instance = response['DBInstances'][0]
                
                instance_class = db_instance.get('DBInstanceClass')
                availability_zone = db_instance.get('AvailabilityZone')
                db_instance_status = db_instance.get('DBInstanceStatus')
                engine_version = db_instance.get('EngineVersion')
                endpoint = db_instance.get('Endpoint', {})
                endpoint_address = endpoint.get('Address')
                multi_az = db_instance.get('MultiAZ')
                

                    
            rds_list.append([rds_id, instance_class, availability_zone, db_instance_status, engine_version, endpoint_address, multi_az  ])
           
            
            sheet.cell(row=row[0].row, column= 2, value=instance_class )
            sheet.cell(row=row[0].row, column= 3, value=availability_zone )
            sheet.cell(row=row[0].row, column= 4, value=db_instance_status )
            sheet.cell(row=row[0].row, column= 5, value=engine_version )
            sheet.cell(row=row[0].row, column= 6, value=endpoint_address )
            sheet.cell(row=row[0].row, column= 6, value=multi_az )
            
            
        except ClientError as e:
              
                
                
                instance_class = 'RDS_Instance Not Found'
                availability_zone = '-'
                db_instance_status = '-'
                engine_version = '-'
                endpoint = '-'
                endpoint_address = '-'
                multi_az = '-'
                
                rds_list.append([rds_id, instance_class, availability_zone, db_instance_status, engine_version, endpoint_address, multi_az  ])
                
            
                sheet.cell(row=row[0].row, column= 2, value=instance_class )
                sheet.cell(row=row[0].row, column= 3, value=availability_zone )
                sheet.cell(row=row[0].row, column= 4, value=db_instance_status )
                sheet.cell(row=row[0].row, column= 5, value=engine_version )
                sheet.cell(row=row[0].row, column= 6, value=endpoint_address )
                sheet.cell(row=row[0].row, column= 6, value=multi_az )
        
        
            
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    rds_table_html ='<div style="margin-bottom: 20px;">'
    rds_table_html +='<h3>RDS Instance Inventory</h3>'
    rds_table_html += "<table border='1'><tr>"
    headers = rds_list[0]
    for header in headers:
        rds_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    rds_table_html += "</tr>"
    
    
    for row in rds_list[1:]:
        rds_table_html += "<tr>"
        for cell in row:
            rds_table_html += f'<td style="background-color: {"rgb(205, 86, 86)"if cell == "RDS_Instance Not Found" else "white"};">{cell}</td>'
        rds_table_html += "</tr>"
    
    rds_table_html += "</table>"  
    rds_table_html += '</div>'
    
    
    
    
######################################################## Code for SG ###########################################

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['SG']
    
    # Find the column index of instance IDs (assuming it's the first column)
    sg_column = 1
    
    # Loop through rows and get Bucket Name
    sg_list = [['SG Id' ,'SG Name', 'SG VPC Id', 'Description' ]]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sg_column, max_col=sg_column):
        sg_id = row[0].value

        try:
            client = boto3.client('ec2')
            
            response = client.describe_security_groups(GroupIds=[sg_id])
            
            name= '-'
            vpc= '-'
            description= '-'

            sg = response['SecurityGroups'][0]
            name= sg.get('GroupName', {})
            vpc= sg.get('VpcId', {})
            description= sg.get('Description', {})
                

                    
            sg_list.append([sg_id,name, vpc,description ])
           
            
            sheet.cell(row=row[0].row, column= 2, value=name )
            sheet.cell(row=row[0].row, column= 3, value=vpc )
            sheet.cell(row=row[0].row, column= 4, value=description )
            
            
        except ClientError as e:
              
                
                name= '-'
                vpc= 'SG_Not Found'
                description= '-'
                
                sg_list.append([sg_id,name, vpc,description ])
                
            
                sheet.cell(row=row[0].row, column= 2, value=name )
                sheet.cell(row=row[0].row, column= 3, value=vpc )
                sheet.cell(row=row[0].row, column= 4, value=description )
                
            
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    sg_table_html ='<div style="margin-bottom: 20px;">'
    sg_table_html +='<h3>Security Group Inventory</h3>'
    sg_table_html += "<table border='1'><tr>"
    headers = sg_list[0]
    for header in headers:
        sg_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    sg_table_html += "</tr>"
    
    
    for row in sg_list[1:]:
        sg_table_html += "<tr>"
        for cell in row:
            sg_table_html += f'<td style="background-color: {"rgb(205, 86, 86)"if cell == "SG_Not Found" else "white"};">{cell}</td>'
        sg_table_html += "</tr>"
    
    sg_table_html += "</table>"  
    sg_table_html += '</div>'
    
    
    
####################################################### Code for RDS Cluster ##########################################

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['DB-Cluster']
    
    # Find the column index of instance IDs (assuming it's the first column)
    db_cluster_column = 1
    
    # Loop through rows and get Bucket Name
    db_cluster_list = [['DB Cluster Id' ,'Status', 'Multi AZ', 'Engine', 'Engine Version', 'Cluster Size', 'Storage' ]]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=db_cluster_column, max_col=db_cluster_column):
        db_cluster_id = row[0].value

        try:
            client = boto3.client('rds')
            
            storage= '-'
            Parametergrp= '-'
            status= '-'
            endpoint= '-'
            readerendpoint = '-'
            MultiAZ= '-'
            Engine= '-'
            EngineVersion= '-'
            cluster_size = '-'
            
            
            response = client.describe_db_clusters(DBClusterIdentifier=db_cluster_id)
            
            storage= response['DBClusters'][0]['AllocatedStorage']
            Parametergrp= response['DBClusters'][0]['DBClusterParameterGroup']
            status= response['DBClusters'][0]['Status']
            endpoint= response['DBClusters'][0]['Endpoint']
            readerendpoint = response['DBClusters'][0]['ReaderEndpoint']
            MultiAZ= response['DBClusters'][0]['MultiAZ']
            Engine= response['DBClusters'][0]['Engine']
            EngineVersion= response['DBClusters'][0]['EngineVersion']
            cluster_size = len(response['DBClusters'][0]['DBClusterMembers'])

            
            

                    
            db_cluster_list.append([db_cluster_id,status, MultiAZ, Engine, EngineVersion, cluster_size, storage ])
           
            
            sheet.cell(row=row[0].row, column= 2, value=status )
            sheet.cell(row=row[0].row, column= 3, value=MultiAZ )
            sheet.cell(row=row[0].row, column= 4, value=Engine )
            sheet.cell(row=row[0].row, column= 5, value=EngineVersion )
            sheet.cell(row=row[0].row, column= 6, value=cluster_size )
            sheet.cell(row=row[0].row, column= 7, value=endpoint )
            sheet.cell(row=row[0].row, column= 8, value=readerendpoint )
            sheet.cell(row=row[0].row, column= 9, value=Parametergrp )
            sheet.cell(row=row[0].row, column= 10, value=storage )
            
            
        except ClientError as e:
              
            db_cluster_id = 'DB Cluster Not Found'
            storage= '-'
            Parametergrp= '-'
            status= '-'
            endpoint= '-'
            readerendpoint = '-'
            MultiAZ= '-'
            Engine= '-'
            EngineVersion= '-'
            cluster_size = '-'
                
            db_cluster_list.append([db_cluster_id,status, MultiAZ, Engine, EngineVersion, cluster_size, storage ])
           
            
            sheet.cell(row=row[0].row, column= 2, value=status )
            sheet.cell(row=row[0].row, column= 3, value=MultiAZ )
            sheet.cell(row=row[0].row, column= 4, value=Engine )
            sheet.cell(row=row[0].row, column= 5, value=EngineVersion )
            sheet.cell(row=row[0].row, column= 6, value=cluster_size )
            sheet.cell(row=row[0].row, column= 7, value=endpoint )
            sheet.cell(row=row[0].row, column= 8, value=readerendpoint )
            sheet.cell(row=row[0].row, column= 9, value=Parametergrp )
            sheet.cell(row=row[0].row, column= 10, value=storage )
            
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    
    db_clu_table_html ='<div style="margin-bottom: 20px;">'
    db_clu_table_html +='<h3>RDS Cluster Inventory</h3>'
    db_clu_table_html += "<table border='1'><tr>"
    headers = db_cluster_list[0]
    for header in headers:
        db_clu_table_html += f"<th style='background-color: {'rgb(201,200,198)'};'>{header}</th>"
    db_clu_table_html += "</tr>"
    
    
    for row in db_cluster_list[1:]:
        db_clu_table_html += "<tr>"
        for cell in row:
            db_clu_table_html += f'<td style="background-color: {"rgb(205, 86, 86)" if cell == "DB Cluster Not Found" else "white"};">{cell}</td>'
        db_clu_table_html += "</tr>"
    
    db_clu_table_html += "</table>"  
    db_clu_table_html += '</div>'
    




#################################################################################################################
    # Upload the Excel file to the S3 bucket
    bucket_name = bucket_name  # Replace with your S3 bucket name
    s3_client = boto3.client('s3')
    s3_client.put_object(Bucket=bucket_name, Key='ec2_inventory.xlsx', Body=excel_file)
    
    
    

    # status_table = tabulate(status_list, headers="firstrow",tablefmt="heavy_grid")
    # print(status_table)
    
    # elb_status_table = tabulate(elb_status_list, headers="firstrow",tablefmt="heavy_grid")
    # print(elb_status_table)
    
    # vol_status_table = tabulate(vol_status_list, headers="firstrow", tablefmt="heavy_grid")
    # print(vol_status_table)
    
    # eip_table = tabulate(eip_list, headers="firstrow", tablefmt="heavy_grid")
    # print(eip_table)
    
    # s3_table = tabulate(s3_list, headers="firstrow", tablefmt="heavy_grid")
    # print(s3_table)
    
    # key_table = tabulate(key_list, headers="firstrow", tablefmt="heavy_grid")
    # print(key_table)
    
    # user_table = tabulate(user_list, headers="firstrow", tablefmt="heavy_grid")
    # print(user_table)
    
    # rds_table = tabulate(rds_list, headers="firstrow", tablefmt="heavy_grid")
    # print(rds_table)
    
    # sg_table = tabulate(sg_list, headers="firstrow", tablefmt="heavy_grid")
    # print(sg_table)
    
    # db_cluster_table = tabulate(db_cluster_list, headers="firstrow", tablefmt="heavy_grid")
    # print(db_cluster_table)
    
    
    
    table_list = [ec2_table_html,elb_table_html,vol_table_html,eip_table_html,s3_table_html,key_table_html,user_table_html,rds_table_html,sg_table_html,db_clu_table_html]
    
    
    tz = pytz.timezone('Asia/Kolkata')
    
    # ##Generate a unique file name with date and timestamp
    
    tz = pytz.timezone('Asia/Kolkata')
    # current_datetime = datetime.datetime.now(tz).strftime("%Y-%m-%d_%H-%M-%S")
    
    current_datetime = datetime.now(tz).strftime("%d-%m-%Y")
    
     # Construct an HTML table using the data from the list
     
     
    final_html = ''
    
    for table_html in table_list:
       if any("<td" in row for row in table_html.splitlines()):
          final_html += table_html
     
    ses_client = boto3.client('ses', region_name= region)  # Replace with your region
    
    sender_email = os.environ.get('sender_email')
    recipient_emails = os.environ.get('recipient_emails').split(',')
    print("Recipient Emails:", recipient_emails)
    to_addresses = [email for email in recipient_emails] 
    print("to_addresses:", to_addresses)
    subject = f"Daily AWS Account Monitoring Report - {current_datetime}"
    
    email_body = f"""
    <html>
    <head></head>
    <body>
    {final_html}
    </body>
    </html>
    """
    
    response = ses_client.send_email(
        Source=sender_email,
        Destination={'ToAddresses': to_addresses}, 
        Message={
            'Subject': {'Data': subject},
            'Body': {
                    'Html': {
                        'Data': (
                                "Good Morning Team..!!<br><br>"
                                f"Account ID: {account}<br>"
                                f"Region: {region}<br>"
                                "<br>"
                                f"{email_body}"
                )
            }
        }
        }
    )
    print("Email sent! Message ID:", response['MessageId'])

    
    

    
    return {
        "statusCode": 200, 
        "body": "Daily Monitoring check completed" 
    }